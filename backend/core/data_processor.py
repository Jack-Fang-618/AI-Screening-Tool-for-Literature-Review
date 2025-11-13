"""
Data Processor - File Parsing and Processing Engine

Handles parsing of Excel, CSV, and RIS files from various databases.
Improved and refined from modern_app.py with better error handling and structure.
"""

import pandas as pd
import rispy
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union
import chardet
import logging
from io import BytesIO, StringIO

logger = logging.getLogger(__name__)


class DataProcessor:
    """
    Unified file processor for systematic review data
    
    Supports:
    - Excel files (.xlsx, .xls) with multiple sheets
    - CSV files with automatic encoding detection
    - RIS files from various databases
    - NBIB files (PubMed/MEDLINE format)
    """
    
    def __init__(self):
        self.supported_extensions = {
            '.xlsx': 'excel',
            '.xls': 'excel_legacy',
            '.csv': 'csv',
            '.ris': 'ris',
            '.txt': 'ris',  # RIS files sometimes have .txt extension
            '.nbib': 'nbib',  # PubMed/MEDLINE format
            '.medline': 'nbib'  # Alternative extension
        }
    
    def detect_file_type(self, file_path: Union[str, Path]) -> str:
        """
        Detect file type based on extension
        
        Args:
            file_path: Path to file
            
        Returns:
            File type: 'excel', 'excel_legacy', 'csv', 'ris', or 'nbib'
            
        Raises:
            ValueError: If file type is not supported
        """
        path = Path(file_path)
        extension = path.suffix.lower()
        
        if extension not in self.supported_extensions:
            raise ValueError(
                f"Unsupported file type: {extension}. "
                f"Supported types: {', '.join(self.supported_extensions.keys())}"
            )
        
        return self.supported_extensions[extension]
    
    def parse_file(
        self,
        file_path: Union[str, Path, BytesIO],
        file_type: Optional[str] = None,
        sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Parse any supported file type into DataFrame
        
        Args:
            file_path: Path to file or BytesIO object
            file_type: Optional file type override
            sheet_name: For Excel files, which sheet to parse
            
        Returns:
            DataFrame with parsed data
            
        Raises:
            Exception: If parsing fails
        """
        # Determine file type
        if file_type is None:
            if isinstance(file_path, (str, Path)):
                file_type = self.detect_file_type(file_path)
            else:
                raise ValueError("file_type must be specified for BytesIO objects")
        
        logger.info(f"Parsing {file_type} file: {file_path}")
        
        # Route to appropriate parser
        if file_type in ['excel', 'excel_legacy']:
            return self.parse_excel(file_path, sheet_name)
        elif file_type == 'csv':
            return self.parse_csv(file_path)
        elif file_type == 'ris':
            return self.parse_ris(file_path)
        elif file_type == 'nbib':
            return self.parse_nbib(file_path)
        else:
            raise ValueError(f"Unknown file type: {file_type}")
    
    def parse_excel(
        self,
        file_path: Union[str, Path, BytesIO],
        sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Parse Excel file
        
        Args:
            file_path: Path to Excel file or BytesIO object
            sheet_name: Which sheet to parse (default: first sheet)
            
        Returns:
            DataFrame with parsed data
        """
        try:
            # Get sheet name if not specified
            if sheet_name is None:
                sheets = self.get_excel_sheets(file_path)
                sheet_name = sheets[0] if sheets else 0
                logger.info(f"Using sheet: {sheet_name}")
            
            # Parse Excel
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Basic cleaning
            df = self._clean_dataframe(df)
            
            logger.info(f"Parsed Excel: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            raise
    
    def parse_csv(
        self,
        file_path: Union[str, Path, BytesIO],
        encoding: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Parse CSV file with automatic encoding detection
        
        Args:
            file_path: Path to CSV file or BytesIO object
            encoding: Optional encoding override
            
        Returns:
            DataFrame with parsed data
        """
        try:
            # Auto-detect encoding if not specified
            if encoding is None and isinstance(file_path, (str, Path)):
                encoding = self._detect_encoding(file_path)
                logger.info(f"Detected encoding: {encoding}")
            elif encoding is None:
                encoding = 'utf-8'  # Default for BytesIO
            
            # Parse CSV
            df = pd.read_csv(file_path, encoding=encoding)
            
            # Basic cleaning
            df = self._clean_dataframe(df)
            
            logger.info(f"Parsed CSV: {len(df)} rows, {len(df.columns)} columns")
            return df
            
        except Exception as e:
            logger.error(f"Failed to parse CSV file: {e}")
            # Try common encodings as fallback
            for fallback_encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    logger.info(f"Trying fallback encoding: {fallback_encoding}")
                    df = pd.read_csv(file_path, encoding=fallback_encoding)
                    df = self._clean_dataframe(df)
                    logger.info(f"Successfully parsed with {fallback_encoding}")
                    return df
                except:
                    continue
            raise
    
    def parse_ris(
        self,
        file_path: Union[str, Path]
    ) -> pd.DataFrame:
        """
        Parse RIS file from various databases
        
        Supports: PubMed, Scopus, Web of Science, Embase, etc.
        
        Args:
            file_path: Path to RIS file
            
        Returns:
            DataFrame with standardized fields
        """
        try:
            # Read RIS file
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
                # Check if file is empty
                if not content.strip():
                    raise ValueError(
                        "RIS file is empty. Please ensure the file contains valid RIS entries."
                    )
                
                # Check for basic RIS format markers
                if not any(marker in content for marker in ['TY  -', 'T1  -', 'TI  -', 'ER  -']):
                    raise ValueError(
                        "File does not appear to be in valid RIS format. "
                        "RIS files should contain tags like 'TY  -', 'TI  -', 'ER  -'. "
                        "Please verify the file format or try exporting again from your database."
                    )
                
                # Reset file pointer and parse
                f.seek(0)
                entries = list(rispy.load(f))
            
            if not entries:
                raise ValueError(
                    "No entries found in RIS file. The file may be corrupted or in an unsupported format. "
                    "Common issues:\n"
                    "1. File encoding might be incorrect (try exporting as UTF-8)\n"
                    "2. RIS tags might be malformed (check for proper spacing after tags)\n"
                    "3. File might have been truncated during download\n"
                    "Try re-exporting the file from your reference database."
                )
            
            logger.info(f"Parsed RIS file: {len(entries)} entries")
            
            # Convert to DataFrame with field mapping
            df = self._ris_to_dataframe(entries)
            
            # Basic cleaning
            df = self._clean_dataframe(df)
            
            return df
            
        except UnicodeDecodeError:
            # Try alternative encodings for RIS
            logger.warning("UTF-8 failed, trying alternative encodings")
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        entries = list(rispy.load(f))
                    df = self._ris_to_dataframe(entries)
                    df = self._clean_dataframe(df)
                    logger.info(f"Successfully parsed with {encoding}")
                    return df
                except:
                    continue
            # If all encodings failed
            raise ValueError(
                "Unable to parse RIS file with any encoding. "
                "The file might be corrupted or in a non-standard format. "
                "Please try:\n"
                "1. Re-exporting from your reference manager with UTF-8 encoding\n"
                "2. Opening the file in a text editor to verify RIS format\n"
                "3. Using a different export format (CSV or Excel)"
            )
        except ValueError:
            # Re-raise ValueError with helpful messages (already formatted above)
            raise
        except Exception as e:
            logger.error(f"Failed to parse RIS file: {e}")
            raise ValueError(
                f"RIS file parsing failed: {str(e)}. "
                "Please ensure the file is a valid RIS export from PubMed, Scopus, Web of Science, or similar databases."
            )
    
    def parse_nbib(
        self,
        file_path: Union[str, Path]
    ) -> pd.DataFrame:
        """
        Parse NBIB (PubMed/MEDLINE) file
        
        NBIB format is PubMed's native export format with fields like:
        - PMID - PubMed ID
        - TI   - Title
        - AB   - Abstract
        - AU   - Author
        - DP   - Publication Date
        - TA   - Journal Title Abbreviation
        - VI   - Volume
        - IP   - Issue
        - PG   - Pagination
        - AID  - Article ID (DOI)
        
        Args:
            file_path: Path to NBIB file
            
        Returns:
            DataFrame with standardized fields
        """
        try:
            # Detect encoding
            encoding = self._detect_encoding(file_path)
            
            # Read file
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            
            # Parse NBIB records
            records = []
            current_record = {}
            current_field = None
            current_value = []
            
            for line in content.split('\n'):
                # Empty line indicates end of record
                if not line.strip():
                    if current_field and current_value:
                        current_record[current_field] = ' '.join(current_value).strip()
                    if current_record:
                        records.append(current_record)
                        current_record = {}
                        current_field = None
                        current_value = []
                    continue
                
                # New field (starts with 4-char tag + space + dash + space)
                if len(line) >= 6 and line[4:6] == '- ':
                    # Save previous field
                    if current_field and current_value:
                        current_record[current_field] = ' '.join(current_value).strip()
                    
                    # Start new field
                    current_field = line[:4].strip()
                    current_value = [line[6:].strip()]
                else:
                    # Continuation of previous field
                    if current_field:
                        current_value.append(line.strip())
            
            # Don't forget last record
            if current_field and current_value:
                current_record[current_field] = ' '.join(current_value).strip()
            if current_record:
                records.append(current_record)
            
            if not records:
                raise ValueError("No records found in NBIB file")
            
            logger.info(f"Parsed NBIB file: {len(records)} records")
            
            # Convert to DataFrame with field mapping
            df = self._nbib_to_dataframe(records)
            
            # Basic cleaning
            df = self._clean_dataframe(df)
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to parse NBIB file: {e}")
            raise
    
    def get_excel_sheets(
        self,
        file_path: Union[str, Path, BytesIO]
    ) -> List[str]:
        """
        Get list of sheet names from Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet names
        """
        try:
            import openpyxl
            if isinstance(file_path, BytesIO):
                wb = openpyxl.load_workbook(file_path, read_only=True)
            else:
                wb = openpyxl.load_workbook(str(file_path), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            logger.warning(f"Could not read sheet names: {e}")
            return []
    
    def _detect_encoding(self, file_path: Union[str, Path]) -> str:
        """
        Detect file encoding using chardet
        
        Args:
            file_path: Path to file
            
        Returns:
            Detected encoding (e.g., 'utf-8', 'latin-1')
        """
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Read first 10KB
            result = chardet.detect(raw_data)
            return result['encoding'] or 'utf-8'
        except Exception as e:
            logger.warning(f"Encoding detection failed: {e}, defaulting to utf-8")
            return 'utf-8'
    
    def _ris_to_dataframe(self, entries: List[Dict]) -> pd.DataFrame:
        """
        Convert RIS entries to DataFrame with standardized field names
        
        RIS field mapping:
        - TI/T1: Title
        - AB/N2: Abstract
        - AU/A1: Authors
        - JO/T2/JF: Journal
        - PY/Y1: Year
        - DO: DOI
        - KW: Keywords
        
        Args:
            entries: List of RIS entry dictionaries
            
        Returns:
            DataFrame with standardized columns
        """
        # Standardized field mapping for RIS
        field_mapping = {
            # Title fields
            'title': ['title', 'primary_title', 'T1', 'TI'],
            # Abstract fields
            'abstract': ['abstract', 'AB', 'N2', 'notes_abstract'],
            # Author fields
            'authors': ['authors', 'AU', 'A1', 'first_authors'],
            # Journal fields
            'journal': ['journal_name', 'secondary_title', 'JO', 'T2', 'JF'],
            # Year fields
            'year': ['year', 'publication_year', 'PY', 'Y1'],
            # DOI fields
            'doi': ['doi', 'DO'],
            # Keywords
            'keywords': ['keywords', 'KW']
        }
        
        # Extract data with flexible field mapping
        data = []
        for entry in entries:
            record = {}
            
            # Apply field mapping
            for std_field, ris_fields in field_mapping.items():
                value = None
                for ris_field in ris_fields:
                    if ris_field in entry and entry[ris_field]:
                        value = entry[ris_field]
                        break
                
                # Handle list values (e.g., authors, keywords)
                if isinstance(value, list):
                    value = '; '.join(str(v) for v in value if v)
                
                record[std_field] = value
            
            data.append(record)
        
        df = pd.DataFrame(data)
        
        # Ensure all standard fields exist
        for field in field_mapping.keys():
            if field not in df.columns:
                df[field] = None
        
        return df
    
    def _nbib_to_dataframe(self, records: List[Dict]) -> pd.DataFrame:
        """
        Convert NBIB records to DataFrame with standardized field names
        
        NBIB field mapping:
        - PMID: PubMed ID
        - TI: Title
        - AB: Abstract
        - AU: Authors
        - FAU: Full Author Name
        - DP: Publication Date
        - TA: Journal Title Abbreviation
        - JT: Journal Title
        - VI: Volume
        - IP: Issue
        - PG: Pagination
        - AID: Article ID (includes DOI)
        - PT: Publication Type
        - MH: MeSH Terms
        - OT: Other Terms (keywords)
        
        Args:
            records: List of NBIB record dictionaries
            
        Returns:
            DataFrame with standardized columns
        """
        data = []
        
        for record in records:
            # Extract fields
            row = {
                'pmid': record.get('PMID', ''),
                'title': record.get('TI', ''),
                'abstract': record.get('AB', ''),
                'authors': record.get('FAU', record.get('AU', '')),  # Prefer full names
                'journal': record.get('JT', record.get('TA', '')),  # Prefer full journal name
                'year': self._extract_year_from_pubmed_date(record.get('DP', '')),
                'volume': record.get('VI', ''),
                'issue': record.get('IP', ''),
                'pages': record.get('PG', ''),
                'publication_type': record.get('PT', ''),
                'mesh_terms': record.get('MH', ''),
                'keywords': record.get('OT', ''),
                'language': record.get('LA', ''),
                'country': record.get('PL', ''),
                'issn': record.get('IS', ''),
                'nlm_id': record.get('JID', ''),
                'pmc_id': record.get('PMC', ''),
            }
            
            # Extract DOI from AID field (format: "10.1234/example [doi]")
            aid = record.get('AID', '')
            if aid:
                if '[doi]' in aid.lower():
                    # Extract DOI (remove [doi] marker)
                    doi = aid.replace('[doi]', '').replace('[DOI]', '').strip()
                    row['doi'] = doi
                else:
                    row['doi'] = aid if '10.' in aid else ''
            else:
                row['doi'] = ''
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        return df
    
    def _extract_year_from_pubmed_date(self, date_str: str) -> str:
        """
        Extract year from PubMed date string
        
        PubMed dates can be in formats like:
        - "2023"
        - "2023 Jan"
        - "2023 Jan 15"
        - "2023 Jan-Feb"
        
        Args:
            date_str: PubMed date string
            
        Returns:
            Year as string, or empty string if not found
        """
        if not date_str:
            return ''
        
        # Extract first 4-digit number (year)
        import re
        match = re.search(r'\b(19|20)\d{2}\b', str(date_str))
        if match:
            return match.group(0)
        
        return ''
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Basic DataFrame cleaning
        
        - Strip whitespace from column names
        - Remove completely empty rows
        - Strip whitespace from string values
        
        Args:
            df: Input DataFrame
            
        Returns:
            Cleaned DataFrame
        """
        # Clean column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Strip whitespace from string columns
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    def validate_dataframe(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate DataFrame for systematic review requirements
        
        Args:
            df: DataFrame to validate
            
        Returns:
            Tuple of (is_valid, list_of_warnings)
        """
        warnings = []
        
        # Check if DataFrame is empty
        if len(df) == 0:
            warnings.append("DataFrame is empty")
            return False, warnings
        
        # Check for required columns (flexible names)
        has_title = any(col.lower() in ['title', 'ti', 'article title', 'primary_title', 't1']
                       for col in df.columns)
        has_abstract = any(col.lower() in ['abstract', 'ab', 'summary', 'n2']
                          for col in df.columns)
        
        if not has_title:
            warnings.append("No title column detected (required for screening)")
        
        if not has_abstract:
            warnings.append("No abstract column detected (recommended for screening)")
        
        # Check for high missing data
        for col in df.columns:
            missing_pct = df[col].isna().sum() / len(df) * 100
            if missing_pct > 90:
                warnings.append(f"Column '{col}' has {missing_pct:.1f}% missing data")
        
        is_valid = has_title  # Only title is strictly required
        
        return is_valid, warnings
    
    def get_file_info(
        self,
        file_path: Union[str, Path],
        file_type: Optional[str] = None
    ) -> Dict:
        """
        Get metadata about a file without fully parsing it
        
        Args:
            file_path: Path to file
            file_type: Optional file type override
            
        Returns:
            Dict with file information
        """
        if file_type is None:
            file_type = self.detect_file_type(file_path)
        
        info = {
            'file_type': file_type,
            'file_size_mb': Path(file_path).stat().st_size / (1024 * 1024)
        }
        
        try:
            if file_type in ['excel', 'excel_legacy']:
                info['sheets'] = self.get_excel_sheets(file_path)
                info['num_sheets'] = len(info['sheets'])
            
            # Get quick row count (first sheet only for Excel)
            df_sample = self.parse_file(file_path, file_type=file_type)
            info['estimated_rows'] = len(df_sample)
            info['columns'] = list(df_sample.columns)
            info['num_columns'] = len(df_sample.columns)
            
        except Exception as e:
            info['error'] = str(e)
        
        return info
